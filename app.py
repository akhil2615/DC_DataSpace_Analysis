from __future__ import annotations

import json
import re
import subprocess
import sys
import threading
import uuid
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from shutil import which
import shutil
from typing import Any

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from starlette.requests import Request

REPO = Path(__file__).resolve().parent
CACHE = REPO / ".data-space-analysis-cache"
RUNS = REPO / "runs" / "web"
SCRIPTS = REPO / "scripts"
FETCH_SCRIPT = SCRIPTS / "fetch_live.py"
FILL_SCRIPT = SCRIPTS / "fill_live.py"

app = FastAPI(title="Data Cloud Data Space Analysis Launcher")
templates = Jinja2Templates(directory=str(REPO / "templates"))

_lock = threading.Lock()
_active_run_id: str | None = None
_runs: dict[str, dict] = {}
_active_compare_id: str | None = None
_compares: dict[str, dict] = {}

COMPAREABLE_ITEMS: dict[str, str] = {
    "data-spaces": "data-spaces.json",
    "data-space-members": "data-space-members.json",
    "data-streams": "data-streams.json",
    "connections": "connections.json",
    "data-lake-objects": "data-lake-objects.json",
    "metadata-dlo": "metadata-dlo.json",
    "metadata-dmo": "metadata-dmo.json",
    "data-model-objects-catalogue": "data-model-objects-catalogue.json",
    "dmo-mappings": "dmo-mappings.json",
    "identity-resolutions": "identity-resolutions.json",
    "segments": "segments.json",
    "activations": "activations.json",
    "activation-targets": "activation-targets.json",
    "activation-details": "activation-details.json",
    "data-transforms": "data-transforms.json",
    "calculated-insights": "calculated-insights.json",
    "metadata-ci": "metadata-ci.json",
    "data-graphs": "data-graphs.json",
    "data-graph-details": "data-graph-details.json",
    "search-index": "search-index.json",
    "data-actions": "data-actions.json",
    "data-action-targets": "data-action-targets.json",
    "connectors-catalog": "connectors-catalog.json",
    "tooling-dmo": "tooling-dmo.json",
    "users": "users.json",
}


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def load_spaces_from_cache() -> list[str]:
    path = CACHE / "data-spaces.json"
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except ValueError:
        return []
    records = payload.get("records")
    if not isinstance(records, list):
        return []
    names = sorted({r.get("name", "").strip() for r in records if isinstance(r, dict) and r.get("name")})
    return [n for n in names if n]


def resolve_sf_cli() -> str | None:
    return which("sf") or which("sf.cmd") or which("sf.exe")


def sf_org_display(target_org: str | None = None, verbose: bool = False) -> dict:
    sf_bin = resolve_sf_cli()
    if not sf_bin:
        return {"ok": False, "error": "Salesforce CLI (sf) was not found in PATH."}
    cmd = [sf_bin, "org", "display"]
    if verbose:
        cmd.append("--verbose")
    cmd.append("--json")
    if target_org:
        cmd.extend(["--target-org", target_org])
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(REPO), check=False)
    except OSError as exc:
        return {"ok": False, "error": f"Failed to run sf CLI: {exc}"}
    if proc.returncode != 0:
        return {"ok": False, "error": (proc.stderr or proc.stdout).strip() or "sf org display failed"}
    try:
        body = json.loads(proc.stdout or "{}")
        res = body.get("result") or {}
    except ValueError:
        return {"ok": False, "error": "Unable to parse sf org display output"}
    return {"ok": True, "result": res}


def read_org_context(target_org: str | None = None) -> dict:
    sf_bin = resolve_sf_cli()
    if not sf_bin:
        return {
            "ok": False,
            "error": "Salesforce CLI (sf) was not found in PATH. Open a terminal where sf works and start the web app from that same environment.",
        }
    primary = sf_org_display(target_org=target_org, verbose=False)
    if not primary.get("ok"):
        return primary
    res = primary.get("result") or {}
    token = str(res.get("accessToken") or "")
    if not token or token.startswith("[REDACTED]"):
        secondary = sf_org_display(target_org=target_org, verbose=True)
        if secondary.get("ok"):
            res = secondary.get("result") or res
    return {
        "ok": True,
        "orgId": res.get("id"),
        "username": res.get("username"),
        "instanceUrl": res.get("instanceUrl"),
        "accessToken": res.get("accessToken"),
        "alias": res.get("alias"),
    }


def check_preflight() -> dict:
    template_path = REPO / "DataCloud_DataSpace_Analysis_Record_TEMPLATE.docx"
    checks = {
        "templateExists": template_path.exists(),
        "fetchScriptExists": FETCH_SCRIPT.exists(),
        "fillScriptExists": FILL_SCRIPT.exists(),
        "cacheExists": CACHE.exists(),
    }
    org = read_org_context()
    checks["cliAuthenticated"] = bool(org.get("ok"))
    return {"checks": checks, "org": org}


def list_authenticated_orgs() -> list[dict]:
    sf_bin = resolve_sf_cli()
    if not sf_bin:
        return []
    cmd = [sf_bin, "org", "list", "--all", "--json"]
    proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(REPO), check=False)
    if proc.returncode != 0:
        return []
    try:
        payload = json.loads(proc.stdout or "{}")
    except ValueError:
        return []
    result = payload.get("result") or {}
    rows = []
    for bucket in ("nonScratchOrgs", "sandboxes", "devHubs", "scratchOrgs"):
        for rec in result.get(bucket) or []:
            if not isinstance(rec, dict):
                continue
            rows.append(
                {
                    "alias": rec.get("alias") or "",
                    "username": rec.get("username") or "",
                    "orgId": rec.get("orgId") or rec.get("id") or "",
                    "isDefault": bool(rec.get("isDefaultUsername") or rec.get("isDefaultDevHubUsername")),
                    "connectedStatus": rec.get("connectedStatus") or rec.get("status") or "",
                }
            )
    # Deduplicate by username/org pair.
    uniq = {}
    for r in rows:
        key = (r["username"], r["orgId"])
        if key not in uniq:
            uniq[key] = r
    return sorted(
        uniq.values(),
        key=lambda x: (
            0 if x["isDefault"] else 1,
            x["alias"] or "~",
            x["username"] or "~",
        ),
    )


def list_spaces_for_org(target_org: str) -> list[str]:
    ctx = read_org_context(target_org=target_org)
    if not ctx.get("ok"):
        raise RuntimeError(ctx.get("error") or "Unable to resolve org context")
    token = str(ctx.get("accessToken") or "")
    if not token or token.startswith("[REDACTED]"):
        raise RuntimeError(
            "CLI token is redacted for this org. Re-login with 'sf org login web --alias <alias>' and retry."
        )
    instance = str(ctx.get("instanceUrl") or "").rstrip("/")
    if not instance:
        raise RuntimeError("instanceUrl missing from sf org display output")
    out: list[str] = []
    offset = 0
    limit = 200
    while True:
        q = urllib.parse.urlencode({"limit": limit, "offset": offset})
        req = urllib.request.Request(
            f"{instance}/services/data/v63.0/ssot/data-spaces?{q}",
            headers={"Authorization": f"Bearer {token}", "Accept": "application/json"},
        )
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                body = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            text = exc.read().decode("utf-8", "replace")
            raise RuntimeError(f"data-spaces query failed ({exc.code}): {text[:220]}") from exc
        records = body.get("dataSpaces") or body.get("records") or []
        rows = [r.get("name", "").strip() for r in records if isinstance(r, dict) and r.get("name")]
        out.extend([r for r in rows if r])
        if len(records) < limit:
            break
        offset += limit
    return sorted(set(out))


def append_log(run_id: str, line: str) -> None:
    with _lock:
        run = _runs.get(run_id)
        if not run:
            return
        run["logs"].append(line.rstrip())
        if len(run["logs"]) > 1000:
            run["logs"] = run["logs"][-1000:]


def load_cache_json(name: str) -> dict:
    p = CACHE / f"{name}.json"
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except ValueError:
        return {}


def excel_safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[:\\/*?\[\]]", "_", name).strip() or "Sheet"
    base = cleaned[:31]
    candidate = base
    suffix = 1
    while candidate in used:
        tail = f"_{suffix}"
        candidate = (base[: 31 - len(tail)] + tail) if len(base) + len(tail) > 31 else (base + tail)
        suffix += 1
    used.add(candidate)
    return candidate


def cell_value(v: Any) -> str | int | float | bool:
    if isinstance(v, (str, int, float, bool)):
        return v
    if v is None:
        return ""
    return json.dumps(v, ensure_ascii=False, separators=(",", ":"))


def write_raw_json_sheet(wb: Workbook, sheet_name: str, payload: Any, used_names: set[str]) -> None:
    ws = wb.create_sheet(excel_safe_sheet_name(sheet_name, used_names))
    ws.append(["json"])
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    chunk_size = 32000  # keep under Excel's 32767 char cell limit
    for i in range(0, len(text), chunk_size):
        ws.append([text[i : i + chunk_size]])


def write_records_sheet(wb: Workbook, sheet_name: str, records: list[dict], used_names: set[str]) -> None:
    ws = wb.create_sheet(excel_safe_sheet_name(sheet_name, used_names))
    if not records:
        ws.append(["info"])
        ws.append(["No records"])
        return
    header: list[str] = []
    seen: set[str] = set()
    for rec in records:
        for key in rec.keys():
            if key not in seen:
                seen.add(key)
                header.append(key)
    ws.append(header)
    for rec in records:
        ws.append([cell_value(rec.get(col)) for col in header])


def build_cache_workbook(run_id: str, space: str, run_dir: Path) -> Path:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Run ID", run_id])
    summary.append(["Data Space", space])
    summary.append(["Generated At UTC", now_utc()])
    summary.append(["Cache Directory", str(CACHE)])

    used_names = {"Summary"}
    cache_files = sorted(CACHE.glob("*.json"))
    summary.append(["Cache File Count", len(cache_files)])

    for f in cache_files:
        summary.append(["Cache File", f.name])
        try:
            payload = json.loads(f.read_text(encoding="utf-8"))
        except ValueError:
            write_raw_json_sheet(wb, f"{f.stem}_raw", {"error": "Invalid JSON"}, used_names)
            continue

        records = payload.get("records") if isinstance(payload, dict) else None
        if isinstance(records, list) and records and all(isinstance(r, dict) for r in records):
            write_records_sheet(wb, f"{f.stem}_records", records, used_names)
        write_raw_json_sheet(wb, f"{f.stem}_raw", payload, used_names)

    out = run_dir / f"DataCloud_DataSpace_Metadata_Bundle_{space}_{run_id}.xlsx"
    wb.save(str(out))
    return out


def run_command(run_id: str, step: str, cmd: list[str]) -> int:
    with _lock:
        _runs[run_id]["step"] = step
    append_log(run_id, f"$ {' '.join(cmd)}")
    try:
        proc = subprocess.Popen(
            cmd,
            cwd=str(REPO),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
        )
    except OSError as exc:
        append_log(run_id, f"[{step}] failed to start: {exc}")
        return 1
    assert proc.stdout is not None
    for line in proc.stdout:
        append_log(run_id, line)
    proc.wait()
    append_log(run_id, f"[{step}] exit_code={proc.returncode}")
    return int(proc.returncode)


def run_compare_command(compare_id: str, step: str, cmd: list[str]) -> int:
    with _lock:
        _compares[compare_id]["step"] = step
    append_compare_log(compare_id, f"$ {' '.join(cmd)}")
    try:
        proc = subprocess.Popen(
            cmd,
            cwd=str(REPO),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
        )
    except OSError as exc:
        append_compare_log(compare_id, f"[{step}] failed to start: {exc}")
        return 1
    assert proc.stdout is not None
    for line in proc.stdout:
        append_compare_log(compare_id, line)
    proc.wait()
    append_compare_log(compare_id, f"[{step}] exit_code={proc.returncode}")
    return int(proc.returncode)


def append_compare_log(compare_id: str, line: str) -> None:
    with _lock:
        run = _compares.get(compare_id)
        if not run:
            return
        run["logs"].append(line.rstrip())
        if len(run["logs"]) > 1500:
            run["logs"] = run["logs"][-1500:]


def record_identity(rec: dict) -> str:
    for k in ("name", "developerName", "apiName", "id", "label"):
        val = rec.get(k)
        if isinstance(val, str) and val.strip():
            return f"{k}:{val.strip()}"
    return "json:" + json.dumps(rec, sort_keys=True, separators=(",", ":"))


def rec_space_name(rec: dict) -> str | None:
    for k in ("dataSpaceName", "dataSpace", "dataspaceName", "dataspace"):
        val = rec.get(k)
        if isinstance(val, str) and val.strip():
            return val.strip()
    return None


def extract_keys(name: str, payload: dict, selected_space: str) -> set[str]:
    if not isinstance(payload, dict):
        return set()
    recs = payload.get("records")
    if isinstance(recs, list):
        out = set()
        for rec in recs:
            if isinstance(rec, dict):
                space_val = rec_space_name(rec)
                if space_val is not None and selected_space and space_val != selected_space:
                    continue
                out.add(record_identity(rec))
            else:
                out.add(f"value:{str(rec)}")
        return out
    if name == "data-space-members":
        by_space = payload.get("bySpace") or {}
        out = set()
        for space, detail in by_space.items():
            if selected_space and space != selected_space:
                continue
            if not isinstance(detail, dict):
                continue
            for m in detail.get("members") or []:
                if isinstance(m, dict):
                    out.add(f"{space}::{m.get('memberName','')}::{m.get('status','')}")
        return out
    if "byDmo" in payload and isinstance(payload["byDmo"], dict):
        return {f"dmo:{k}" for k in payload["byDmo"].keys()}
    if "byType" in payload and isinstance(payload["byType"], dict):
        out = set()
        for ctype, rows in payload["byType"].items():
            if not isinstance(rows, list):
                continue
            for rec in rows:
                if isinstance(rec, dict):
                    out.add(f"{ctype}::{record_identity(rec)}")
        return out
    if "byActivation" in payload and isinstance(payload["byActivation"], dict):
        return {f"activation:{k}" for k in payload["byActivation"].keys()}
    if "byGraph" in payload and isinstance(payload["byGraph"], dict):
        return {f"graph:{k}" for k in payload["byGraph"].keys()}
    if "byId" in payload and isinstance(payload["byId"], dict):
        return {f"user:{k}" for k in payload["byId"].keys()}
    return set()


def build_compare_rows(
    source_cache: Path,
    dest_cache: Path,
    source_space: str,
    destination_space: str,
    selected_items: list[str] | None = None,
) -> list[dict]:
    source_files = {p.name for p in source_cache.glob("*.json")}
    dest_files = {p.name for p in dest_cache.glob("*.json")}
    if selected_items:
        chosen_files = {
            COMPAREABLE_ITEMS[item]
            for item in selected_items
            if item in COMPAREABLE_ITEMS
        }
        files = sorted(chosen_files)
    else:
        files = sorted((source_files | dest_files) - {"_provenance.json"})
    rows: list[dict] = []
    for name in files:
        s_payload = {}
        d_payload = {}
        s_path = source_cache / name
        d_path = dest_cache / name
        if s_path.exists():
            try:
                s_payload = json.loads(s_path.read_text(encoding="utf-8"))
            except ValueError:
                s_payload = {}
        if d_path.exists():
            try:
                d_payload = json.loads(d_path.read_text(encoding="utf-8"))
            except ValueError:
                d_payload = {}
        s_keys = extract_keys(name.replace(".json", ""), s_payload, source_space)
        d_keys = extract_keys(name.replace(".json", ""), d_payload, destination_space)
        missing_in_dest = sorted(s_keys - d_keys)
        missing_in_source = sorted(d_keys - s_keys)
        rows.append(
            {
                "item": name.replace(".json", ""),
                "sourceCount": len(s_keys),
                "destinationCount": len(d_keys),
                "missingInDestinationCount": len(missing_in_dest),
                "missingInSourceCount": len(missing_in_source),
                "missingInDestination": missing_in_dest,
                "missingInSource": missing_in_source,
            }
        )
    return rows


def start_run(space: str, fresh_fetch: bool) -> str:
    global _active_run_id
    safe_space = re.sub(r"[^A-Za-z0-9_.-]+", "_", space)
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ") + "-" + uuid.uuid4().hex[:8]
    run_dir = RUNS / run_id
    docs_dir = run_dir / "docs"
    docs_dir.mkdir(parents=True, exist_ok=True)

    with _lock:
        if _active_run_id:
            raise RuntimeError("Another run is currently in progress.")
        _active_run_id = run_id
        _runs[run_id] = {
            "id": run_id,
            "space": space,
            "safeSpace": safe_space,
            "status": "running",
            "step": "queued",
            "freshFetch": fresh_fetch,
            "logs": [],
            "startedAtUtc": now_utc(),
            "finishedAtUtc": None,
            "outputFile": None,
            "workbookFile": None,
            "runDir": str(run_dir),
            "error": None,
        }

    def worker() -> None:
        global _active_run_id
        try:
            preflight = check_preflight()
            checks = preflight["checks"]
            if not all((checks["templateExists"], checks["fetchScriptExists"], checks["fillScriptExists"], checks["cliAuthenticated"])):
                raise RuntimeError(f"Preflight failed: {json.dumps(preflight, separators=(',', ':'))}")

            if fresh_fetch:
                rc = run_command(
                    run_id,
                    "fetch",
                    [
                        sys.executable,
                        "-u",
                        str(FETCH_SCRIPT),
                        "--clean-cache",
                        "--workers",
                        "8",
                        "--max-retries",
                        "4",
                        "--retry-base-ms",
                        "400",
                        "--adaptive-throttle",
                    ],
                )
                if rc != 0:
                    raise RuntimeError("Fetch step failed. See logs.")

            rc = run_command(
                run_id,
                "fill",
                [
                    sys.executable,
                    "-u",
                    str(FILL_SCRIPT),
                    "--space",
                    space,
                    "--output-dir",
                    str(docs_dir),
                ],
            )
            if rc != 0:
                raise RuntimeError("Fill step failed. See logs.")

            out_candidates = sorted(docs_dir.glob(f"DataCloud_DataSpace_Analysis_Record_{safe_space}_LIVE_*.docx"))
            output_file = str(out_candidates[-1]) if out_candidates else None
            if output_file:
                source = Path(output_file)
                renamed = docs_dir / f"DataCloud_DataSpace_Analysis_Record_{safe_space}_{run_id}.docx"
                shutil.copy2(source, renamed)
                output_file = str(renamed)
            workbook_file = str(build_cache_workbook(run_id=run_id, space=safe_space, run_dir=run_dir))

            with _lock:
                run = _runs[run_id]
                run["status"] = "completed"
                run["step"] = "done"
                run["finishedAtUtc"] = now_utc()
                run["outputFile"] = output_file
                run["workbookFile"] = workbook_file
                if not output_file:
                    run["error"] = "Run finished but output file was not found."
                    run["status"] = "failed"
        except Exception as exc:
            with _lock:
                run = _runs[run_id]
                run["status"] = "failed"
                run["step"] = "error"
                run["finishedAtUtc"] = now_utc()
                run["error"] = str(exc)
                run["logs"].append(f"[error] {exc}")
        finally:
            with _lock:
                _active_run_id = None

    threading.Thread(target=worker, daemon=True).start()
    return run_id


def start_compare(
    source_org: str,
    source_space: str,
    destination_org: str,
    destination_space: str,
    selected_items: list[str] | None = None,
) -> str:
    global _active_compare_id
    compare_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ") + "-cmp-" + uuid.uuid4().hex[:8]
    compare_dir = RUNS / compare_id
    source_cache = compare_dir / "source-cache"
    destination_cache = compare_dir / "destination-cache"
    compare_dir.mkdir(parents=True, exist_ok=True)

    with _lock:
        if _active_compare_id:
            raise RuntimeError("Another comparison is currently in progress.")
        _active_compare_id = compare_id
        _compares[compare_id] = {
            "id": compare_id,
            "status": "running",
            "step": "queued",
            "startedAtUtc": now_utc(),
            "finishedAtUtc": None,
            "sourceOrg": source_org,
            "sourceSpace": source_space,
            "destinationOrg": destination_org,
            "destinationSpace": destination_space,
            "selectedItems": selected_items or [],
            "logs": [],
            "rows": [],
            "outputFile": None,
            "error": None,
        }

    def worker() -> None:
        global _active_compare_id
        try:
            preflight = check_preflight()
            checks = preflight["checks"]
            if not all((checks["fetchScriptExists"], checks["cliAuthenticated"])):
                raise RuntimeError(f"Preflight failed: {json.dumps(preflight, separators=(',', ':'))}")

            common_args = [
                "--clean-cache",
                "--workers",
                "8",
                "--max-retries",
                "4",
                "--retry-base-ms",
                "400",
                "--adaptive-throttle",
            ]

            rc = run_compare_command(
                compare_id,
                "fetch_source",
                [
                    sys.executable,
                    "-u",
                    str(FETCH_SCRIPT),
                    "--target-org",
                    source_org,
                    "--cache-dir",
                    str(source_cache),
                    *common_args,
                ],
            )
            if rc != 0:
                raise RuntimeError("Source fetch failed. See comparison logs.")

            rc = run_compare_command(
                compare_id,
                "fetch_destination",
                [
                    sys.executable,
                    "-u",
                    str(FETCH_SCRIPT),
                    "--target-org",
                    destination_org,
                    "--cache-dir",
                    str(destination_cache),
                    *common_args,
                ],
            )
            if rc != 0:
                raise RuntimeError("Destination fetch failed. See comparison logs.")

            with _lock:
                _compares[compare_id]["step"] = "compare"
            rows = build_compare_rows(
                source_cache=source_cache,
                dest_cache=destination_cache,
                source_space=source_space,
                destination_space=destination_space,
                selected_items=selected_items,
            )
            output_path = compare_dir / "comparison.json"
            output_path.write_text(json.dumps({"rows": rows}, indent=2), encoding="utf-8")
            with _lock:
                run = _compares[compare_id]
                run["rows"] = rows
                run["outputFile"] = str(output_path)
                run["status"] = "completed"
                run["step"] = "done"
                run["finishedAtUtc"] = now_utc()
        except Exception as exc:
            with _lock:
                run = _compares[compare_id]
                run["status"] = "failed"
                run["step"] = "error"
                run["finishedAtUtc"] = now_utc()
                run["error"] = str(exc)
                run["logs"].append(f"[error] {exc}")
        finally:
            with _lock:
                _active_compare_id = None

    threading.Thread(target=worker, daemon=True).start()
    return compare_id


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    # Use keyword arguments for compatibility across Starlette template APIs.
    return templates.TemplateResponse(request=request, name="index.html", context={})


@app.get("/api/preflight")
def api_preflight():
    return check_preflight()


@app.get("/api/spaces")
def api_spaces():
    spaces = load_spaces_from_cache()
    return {"spaces": spaces, "count": len(spaces)}


@app.get("/api/orgs")
def api_orgs():
    rows = list_authenticated_orgs()
    return {"orgs": rows, "count": len(rows)}


@app.get("/api/org-spaces")
def api_org_spaces(org: str):
    org = (org or "").strip()
    if not org:
        raise HTTPException(status_code=400, detail="org is required")
    try:
        spaces = list_spaces_for_org(org)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return {"org": org, "spaces": spaces, "count": len(spaces)}


@app.get("/api/compare/items")
def api_compare_items():
    items = [{"id": k, "file": v} for k, v in sorted(COMPAREABLE_ITEMS.items())]
    return {"items": items, "count": len(items)}


@app.post("/api/spaces/refresh")
def api_spaces_refresh():
    with _lock:
        if _active_run_id:
            raise HTTPException(status_code=409, detail="Cannot refresh while a run is active.")
    cmd = [
        sys.executable,
        "-u",
        str(FETCH_SCRIPT),
        "--clean-cache",
        "--workers",
        "8",
        "--max-retries",
        "4",
        "--retry-base-ms",
        "400",
        "--adaptive-throttle",
    ]
    proc = subprocess.run(cmd, cwd=str(REPO), capture_output=True, text=True, check=False)
    if proc.returncode != 0:
        raise HTTPException(status_code=500, detail=(proc.stdout or "") + "\n" + (proc.stderr or ""))
    return {"ok": True, "spaces": load_spaces_from_cache()}


@app.post("/api/run")
def api_run(payload: dict):
    space = str(payload.get("space", "")).strip()
    fresh_fetch = bool(payload.get("freshFetch", True))
    if not space:
        raise HTTPException(status_code=400, detail="space is required")
    if space not in load_spaces_from_cache() and not fresh_fetch:
        raise HTTPException(status_code=400, detail="Space not found in cache. Use freshFetch or refresh spaces first.")
    try:
        run_id = start_run(space=space, fresh_fetch=fresh_fetch)
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return {"runId": run_id}


@app.get("/api/run/active")
def api_run_active():
    with _lock:
        return {"runId": _active_run_id}


@app.get("/api/compare/active")
def api_compare_active():
    with _lock:
        return {"compareId": _active_compare_id}


@app.get("/api/runs/{run_id}")
def api_run_status(run_id: str):
    with _lock:
        run = _runs.get(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    return run


@app.post("/api/compare")
def api_compare(payload: dict):
    source_org = str(payload.get("sourceOrg", "")).strip()
    source_space = str(payload.get("sourceSpace", "")).strip()
    destination_org = str(payload.get("destinationOrg", "")).strip()
    destination_space = str(payload.get("destinationSpace", "")).strip()
    selected_items_raw = payload.get("selectedItems", [])
    if not isinstance(selected_items_raw, list):
        raise HTTPException(status_code=400, detail="selectedItems must be a list")
    selected_items = [str(x).strip() for x in selected_items_raw if str(x).strip()]
    invalid = [x for x in selected_items if x not in COMPAREABLE_ITEMS]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Unknown compare items: {', '.join(invalid)}")
    if not all((source_org, source_space, destination_org, destination_space)):
        raise HTTPException(
            status_code=400,
            detail="sourceOrg, sourceSpace, destinationOrg and destinationSpace are required",
        )
    try:
        compare_id = start_compare(
            source_org=source_org,
            source_space=source_space,
            destination_org=destination_org,
            destination_space=destination_space,
            selected_items=selected_items,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return {"compareId": compare_id}


@app.get("/api/compares/{compare_id}")
def api_compare_status(compare_id: str):
    with _lock:
        run = _compares.get(compare_id)
    if not run:
        raise HTTPException(status_code=404, detail="Comparison not found")
    return run


@app.get("/api/compares/{compare_id}/download")
def api_compare_download(compare_id: str):
    with _lock:
        run = _compares.get(compare_id)
    if not run:
        raise HTTPException(status_code=404, detail="Comparison not found")
    output_file = run.get("outputFile")
    if not output_file:
        raise HTTPException(status_code=404, detail="No comparison output file for this run")
    path = Path(output_file)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Comparison output path no longer exists")
    return FileResponse(path, filename=path.name, media_type="application/json")


@app.get("/api/runs/{run_id}/download")
def api_run_download(run_id: str):
    with _lock:
        run = _runs.get(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    output_file = run.get("outputFile")
    if not output_file:
        raise HTTPException(status_code=404, detail="No output file for this run")
    path = Path(output_file)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Output file path no longer exists")
    return FileResponse(path, filename=path.name, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@app.get("/api/runs/{run_id}/download-workbook")
def api_run_download_workbook(run_id: str):
    with _lock:
        run = _runs.get(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    workbook_file = run.get("workbookFile")
    if not workbook_file:
        raise HTTPException(status_code=404, detail="No workbook file for this run")
    path = Path(workbook_file)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Workbook file path no longer exists")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _html_escape(text: str) -> str:
    return (
        (text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def docx_to_html(path: Path) -> str:
    """Render a generated .docx into simple, readable HTML for in-browser preview."""
    import docx
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    doc = docx.Document(str(path))
    parts: list[str] = []
    open_list = False

    def render_paragraph(par) -> str:
        nonlocal open_list
        text = _html_escape(par.text)
        style = (par.style.name if par.style else "") or ""
        style_l = style.lower()
        if style_l.startswith("list"):
            prefix = "" if open_list else "<ul>"
            open_list = True
            return f"{prefix}<li>{text or '&nbsp;'}</li>"
        closing = "</ul>" if open_list else ""
        open_list = False
        if not text.strip():
            return closing
        if style_l == "title":
            return f"{closing}<h1>{text}</h1>"
        if style_l.startswith("heading 1"):
            return f"{closing}<h2>{text}</h2>"
        if style_l.startswith("heading 2"):
            return f"{closing}<h3>{text}</h3>"
        if style_l.startswith("heading 3"):
            return f"{closing}<h4>{text}</h4>"
        return f"{closing}<p>{text}</p>"

    def render_table(tbl) -> str:
        nonlocal open_list
        closing = "</ul>" if open_list else ""
        open_list = False
        rows_html = []
        for i, row in enumerate(tbl.rows):
            tag = "th" if i == 0 else "td"
            cells = "".join(
                f"<{tag}>{_html_escape(c.text).replace(chr(10), '<br/>')}</{tag}>"
                for c in row.cells
            )
            rows_html.append(f"<tr>{cells}</tr>")
        return f'{closing}<table class="doc-table">{"".join(rows_html)}</table>'

    for child in doc.element.body.iterchildren():
        tag = child.tag
        if tag.endswith("}p"):
            parts.append(render_paragraph(Paragraph(child, doc)))
        elif tag.endswith("}tbl"):
            parts.append(render_table(Table(child, doc)))
    if open_list:
        parts.append("</ul>")
    return "\n".join(p for p in parts if p)


def xlsx_to_preview(path: Path, max_rows: int = 200) -> list[dict]:
    """Return each worksheet as a capped list of rows for in-browser preview."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets: list[dict] = []
    try:
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            total = 0
            for r in ws.iter_rows(values_only=True):
                total += 1
                if len(rows) < max_rows:
                    rows.append(["" if v is None else str(v) for v in r])
            sheets.append(
                {
                    "name": ws.title,
                    "rows": rows,
                    "totalRows": total,
                    "truncated": total > max_rows,
                }
            )
    finally:
        wb.close()
    return sheets


@app.get("/api/runs/{run_id}/preview")
def api_run_preview(run_id: str):
    with _lock:
        run = _runs.get(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    result: dict[str, Any] = {
        "docxHtml": None,
        "docName": None,
        "sheets": None,
        "workbookName": None,
    }
    output_file = run.get("outputFile")
    if output_file and Path(output_file).exists():
        try:
            result["docxHtml"] = docx_to_html(Path(output_file))
            result["docName"] = Path(output_file).name
        except Exception as exc:  # noqa: BLE001 - preview should never hard-fail
            result["docxHtml"] = f"<p>Could not render document preview: {_html_escape(str(exc))}</p>"
            result["docName"] = Path(output_file).name
    workbook_file = run.get("workbookFile")
    if workbook_file and Path(workbook_file).exists():
        try:
            result["sheets"] = xlsx_to_preview(Path(workbook_file))
            result["workbookName"] = Path(workbook_file).name
        except Exception as exc:  # noqa: BLE001
            result["sheets"] = [
                {"name": "error", "rows": [[f"Could not render workbook preview: {exc}"]], "totalRows": 1, "truncated": False}
            ]
            result["workbookName"] = Path(workbook_file).name
    if result["docxHtml"] is None and result["sheets"] is None:
        raise HTTPException(status_code=404, detail="No generated files available to preview")
    return result
